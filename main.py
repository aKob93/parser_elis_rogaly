# -*- coding: utf8 -*-
import shutil
import os.path
import time

from fake_useragent import UserAgent
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from tqdm import tqdm
import datetime
from PIL import Image
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver import Keys

TOKEN = ''
SECRET_KEY = ''
ua = UserAgent()
headers = {'user_agent': ua.random}

read_file = ''
for file in os.listdir():
    if file[:5] == 'data.':
        read_file = file


def open_token_file():
    try:
        with open('token.txt', 'r') as file:
            for i, line in enumerate(file):
                if i == 0:
                    global TOKEN
                    TOKEN = line.split('=')[1].strip().split(', ')
                elif i == 1:
                    global SECRET_KEY
                    SECRET_KEY = line.split('=')[1].strip().split(', ')

    except Exception:
        print('Не удалось прочитать token или secret_key')
        raise IndexError


def get_article_number(read_file):
    try:
        wb = load_workbook(filename=read_file)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        article_numbers = []

        for row in ws.iter_cols(min_col=5, max_col=5, min_row=13):
            for cell in row:
                if cell.value is None:
                    continue
                article_numbers.append(cell.value)
        return article_numbers
    except Exception as exc:
        print(f'Ошибка {exc} в чтении табличного документа data.xlsm')
        with open('error.txt', 'a', encoding='utf-8') as file:

            file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                       f'Ошибка {exc} в чтении табличного документа data.xlsm, функция - get_article_number()\n')
        raise IndexError


def get_link_product_emulation_browser(article_numbers):
    try:
        links_products = []

        options_chrome = webdriver.ChromeOptions()

        options_chrome.add_argument('--headless')  # запуск браузера без в скрытом режиме
        with webdriver.Chrome(options=options_chrome, ) as browser:
            # browser.set_window_position(9999999, 99999999)
            browser.get('https://elis.ru/')
            # browser.maximize_window()
            WebDriverWait(browser, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="search-open"]'))).click()
            time.sleep(1)
            for art in tqdm(article_numbers):
                try:
                    ActionChains(browser).send_keys(art).perform()
                    ActionChains(browser).send_keys(Keys.SPACE).perform()
                    WebDriverWait(browser, 30).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="search-wrap"]/div/div/div/div[1]/a')))
                    time.sleep(2)
                    tt = browser.find_element(By.XPATH, '//*[@id="search-wrap"]/div/div').find_elements(By.CLASS_NAME,
                                                                                                        'card__image')
                    for link in tt:
                        links_products.append(link.find_element(By.TAG_NAME, 'a').get_attribute('href'))
                    ActionChains(browser).key_down(Keys.SHIFT).send_keys(Keys.HOME).key_up(Keys.SHIFT).perform()
                    ActionChains(browser).send_keys(Keys.BACKSPACE).perform()
                except Exception:
                    ActionChains(browser).key_down(Keys.SHIFT).send_keys(Keys.HOME).key_up(Keys.SHIFT).perform()
                    ActionChains(browser).send_keys(Keys.BACKSPACE).perform()
                    with open('erro_article.txt', 'a', encoding='utf-8') as file:
                        file.write(f'Не найден товар - {art}\n')
            for x in links_products:
                if x == ' ':
                    continue
                elif links_products.count(x) > 1:
                    links_products.remove(x)
            return links_products
    except Exception as exc:
        print(f'Ошибка {exc} в получении ссылок на товары')
        with open('error.txt', 'a', encoding='utf-8') as file:

            file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                       f'Ошибка {exc} в получении ссылок на товары, функция - get_link_product()\n')
        raise IndexError


# def get_link_product(article_numbers):
#     links_products = []
#
#     try:
#         for article in tqdm(article_numbers):
#             url = f'https://elis.ru/search/?q={article}'
#             response = requests.get(url, headers=headers)
#             soup = BeautifulSoup(response.text, features='html.parser')
#             product_not_found = soup.find('p', class_='catalog__query').text
#
#             if 'ничего не найдено' in product_not_found:
#                 print(product_not_found)
#                 continue
#
#             found_links_products = soup.find('div', class_='catalog__grid').find_all(class_='js-ga-link-click')
#
#             for link in found_links_products:
#                 links_products.append(link["href"])
#
#         # убираются дубликаты артикулов
#         for x in links_products:
#             if x == ' ':
#                 continue
#             elif links_products.count(x) > 1:
#                 links_products.remove(x)
#         return links_products
#     except Exception as exc:
#         print(f'Ошибка {exc} в получении ссылок на товары')
#         with open('error.txt', 'a', encoding='utf-8') as file:
#
#             file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
#                        f'Ошибка {exc} в получении ссылок на товары, функция - get_link_product()\n')
#         raise IndexError


def get_link_img(links_products):
    article_imgs = {}
    try:
        for line in tqdm(links_products):
            if line == " ":
                continue
            response = requests.get(f'{line.rstrip()}', headers=headers)

            soup = BeautifulSoup(response.text, features='html.parser')

            found_links_imgs = soup.find('div', class_='item-detail-list swiper-wrapper').find_all('a')
            if len(found_links_imgs) < 4:
                link_image = [found_links_imgs[0].find('img')['src'], found_links_imgs[1].find('img')['src'],
                              found_links_imgs[2].find('img')['src']]
            else:
                link_image = [found_links_imgs[0].find('img')['src'], found_links_imgs[1].find('img')['src'],
                              found_links_imgs[2].find('img')['src'], found_links_imgs[3].find('img')['src']]

            color = soup.find('div', class_='ac-cat-list__sub').find_all('p')
            article = soup.find('div', class_='item-detail__descr').find(class_='art')
            #  описание товара
            try:
                description = soup.find('div', class_='item-detail__detail').find('p').text
            except Exception:
                description = ''
            # характеристики товара
            try:
                specifications_found = soup.find_all('p', class_='item-info')
                list_specifications_found = [spec.text.strip() for spec in specifications_found]
                specifications = '\n'.join(list_specifications_found)
            except Exception:
                specifications = ''

            article_imgs[line] = {'Артикул': article.text.split('арт: ')[1].rstrip(),
                                  'Цвет': color[0].text.split('Цвет: ')[1].rstrip(),
                                  'Описание': description,
                                  'Характеристики': specifications,
                                  'Картинка': link_image}
        return article_imgs
    except Exception as exc:
        print(f'Ошибка {exc} в получении ссылок на изображения товаров')
        with open('error.txt', 'a', encoding='utf-8') as file:

            file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                       f'Ошибка {exc} в получении ссылок на изображения товаров, функция - get_link_img()\n')
        raise IndexError


def save_image(article_imgs):
    try:
        if not os.path.isdir('./img/'):
            os.mkdir('./img/')
        for key in article_imgs:
            images = []
            for a, img in enumerate(article_imgs[key]['Картинка']):
                res = requests.get(f'https://elis.ru{img}', stream=True)
                if res.status_code == 200:
                    date_now = datetime.datetime.now()
                    with open(f'./img/{article_imgs[key]["Артикул"]}_{date_now.strftime("%M%S%f")}_{a}.jpg', 'wb') as f:
                        shutil.copyfileobj(res.raw, f)
                        images.append(f'./img/{article_imgs[key]["Артикул"]}_{date_now.strftime("%M%S%f")}_{a}.jpg')
                    print(
                        f'Изображение сохранено - {article_imgs[key]["Артикул"]}_{date_now.strftime("%M%S%f")}_{a}.jpg')
                else:
                    print(f'Изображение {article_imgs[key]["Артикул"]} не удалось сохранить ')
            article_imgs[key]['Картинка'] = images
        return article_imgs
    except Exception as exc:
        print(f'Ошибка {exc} в скачивании изображений')
        with open('error.txt', 'a', encoding='utf-8') as file:

            file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                       f'Ошибка {exc} в скачивании изображений, функция - save_image()\n')
        raise IndexError


def resize_img():
    try:
        for img_file in tqdm(os.listdir('./img/')):
            if img_file[-4:] == '.jpg':
                img = Image.open(f'./img/{img_file}')
                new_image = img.resize((320, 426))
                new_image.save(f'./img/{img_file}')
    except Exception as exc:
        print(f'Ошибка {exc} в изменении разрешения изображений')
        with open('error.txt', 'a', encoding='utf-8') as file:
            file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                       f'Ошибка {exc} в изменении разрешения изображений, функция - resize_img()\n')
        raise IndexError


def sending_to_fotohosting(token, secret_key, images_url):
    active_token = TOKEN[0]
    active_secret_key = SECRET_KEY[0]
    headers = {
        'Authorization': f'TOKEN {active_token}',
    }
    for img_url in images_url:
        img_short_link = []
        print(f'\rЗагрузка изображений товара - {images_url[img_url]["Артикул"]}')
        img_links = images_url[img_url]['Картинка']
        for img in tqdm(img_links):
            try:
                files = {
                    'image': open(img, 'rb'),
                    'secret_key': (None, active_secret_key),
                }
                response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                if response.json()['status'] == 200:
                    img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                          f"[/IMG][/URL]")
                else:
                    print(f'не удалось загрузить {img}')
                    continue
            except KeyError:
                print(f'{img_url} ошибка загрузки изображения - {response.json()["error"]["message"]}\n')
                with open('error.txt', 'a', encoding='utf-8') as file:
                    file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                               f'{img} ошибка загрузки изображения, функция - sending_to_fotohosting()\n')
                if response.json()["error"]["message"] == 'File reception error':
                    continue
                elif response.json()["error"]["message"] == \
                        'Exceeded the daily limit of uploaded images for your account':
                    print('Переключение на второй аккаунт')

                    active_token = TOKEN[1]
                    active_secret_key = SECRET_KEY[1]
                    headers = {
                        'Authorization': f'TOKEN {active_token}',
                    }
                    files = {
                        'image': open(img, 'rb'),
                        'secret_key': (None, active_secret_key),
                    }
                    response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                    if response.json()['status'] == 200:
                        img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                              f"[/IMG][/URL]")
                    else:
                        print(f'Не удалось загрузить {img}')
                continue
            images_url[img_url]['Картинка'] = img_short_link
    return images_url


def write_final_file(article_and_short_links_imgs, read_file):
    try:
        columns = ['X', 'Y', 'Z', 'AA']

        wb = load_workbook(filename=read_file)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        ws['V10'] = 'Описание'
        ws['W10'] = 'Характеристики'
        ws['X10'] = 'Ссылки на фотографии'
        date_now = datetime.datetime.now()
        # запись в файл изображений
        for article in article_and_short_links_imgs.keys():
            for a, link in enumerate(article_and_short_links_imgs[article]['Картинка']):
                for row in ws.iter_cols(min_col=5, max_col=7, min_row=13):
                    for cell in row:
                        if article_and_short_links_imgs[article]['Артикул'] == ws[f'E{cell.row}'].value and \
                                article_and_short_links_imgs[article]['Цвет'] == ws[f'G{cell.row}'].value:
                            ws[f'{columns[a]}{cell.row}'] = link

        # запись в файл хар-ик и описание
        for article in article_and_short_links_imgs.keys():
            for row in ws.iter_cols(min_col=20, max_col=20, min_row=13):
                for cell in row:
                    if article_and_short_links_imgs[article]['Артикул'] == ws[f'E{cell.row}'].value and \
                            article_and_short_links_imgs[article]['Цвет'] == ws[f'G{cell.row}'].value:
                        ws[f'V{cell.row}'] = article_and_short_links_imgs[article]['Описание']
                        ws[f'W{cell.row}'] = article_and_short_links_imgs[article]['Характеристики']

        file_name = f'data_final_{date_now.strftime("%d-%m-%y_%H-%M")}.xlsx'
        wb.save(filename=file_name)
        shutil.rmtree('./img/')
        print(f'Файл {file_name} сохранён')
    except Exception as exc:
        print(f'Ошибка {exc} в записи итогового файла')
        with open('error.txt', 'a', encoding='utf-8') as file:
            file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                       f'Ошибка {exc} в записи итогового файла, функция - write_final_file()\n')
        raise IndexError


def main():
    try:
        print('Начало работы')
        open_token_file()
        print('Сбор данных с таблицы')
        article_numbers = get_article_number(read_file)
        # get_link_product(article_numbers)
        print('\rСбор завершён')
        print('---------------------------\n')
        print('Сбор ссылок на товары')

        links_products = get_link_product_emulation_browser(article_numbers)
        # links_products = get_link_product(article_numbers)
        print('\rСбор завершён')
        print('---------------------------\n')
        print('Сбор изображений с товаров')
        article_imgs = get_link_img(links_products)
        print('\rСбор завершён')
        print('---------------------------\n')
        print('Скачивание изображений')
        article_save_imgs = save_image(article_imgs)
        print('\rСкачивание завершено')
        print('---------------------------\n')
        print('Изменение разрешения изображений')
        resize_img()
        print('\rИзменение завершено')
        print('---------------------------\n')
        print('Загрузка изображений на фотохостинг')
        article_and_short_links_imgs = sending_to_fotohosting(TOKEN, SECRET_KEY, article_save_imgs)
        print('\rЗагрузка завершена')
        print('---------------------------\n')
        print('Запись в файл')
        write_final_file(article_and_short_links_imgs, read_file)
        print('Работа завершена')
        print('Для выхода нажмите Enter')
        input()
        print('---------------------------\n')
    except Exception as exc:
        print(f'Произошла ошибка {exc}')
        print('Для выхода нажмите Enter')
        input()
        print('---------------------------\n')


if __name__ == '__main__':
    main()
